import openpyxl
import re

file_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT).xlsx'
wb = openpyxl.load_workbook(file_path)

# Step 1: Load affordability sheet and create food code mapping
affordability_sheet = wb['Affordability']
food_code_map = {}

print("Loading Affordability data...")
for row in range(2, affordability_sheet.max_row + 1):
    code = affordability_sheet.cell(row, 1).value  # food_code
    description = affordability_sheet.cell(row, 2).value  # food_description
    if code and description:
        # Normalize the description for matching
        normalized = str(description).lower().strip()
        food_code_map[normalized] = code

print(f"Loaded {len(food_code_map)} food items from Affordability sheet\n")

def normalize_food_name(name):
    """Normalize food name for matching"""
    if not name:
        return ""
    # Convert to lowercase, remove extra spaces, remove special chars except spaces and hyphens
    name = str(name).lower().strip()
    name = re.sub(r'[^\w\s\-]', '', name)  # Remove special characters except letters, numbers, spaces, hyphens
    name = re.sub(r'\s+', ' ', name)  # Replace multiple spaces with single space
    return name

def quick_match(food_name, food_code_map):
    """Find matching food code using normalized string matching"""
    if not food_name:
        return None
    
    normalized_name = normalize_food_name(food_name)
    
    # Try exact match with normalized names
    for mapped_name, code in food_code_map.items():
        if normalize_food_name(mapped_name) == normalized_name:
            return code
    
    # Try substring matching (if food name contains most of the key words)
    name_words = set(normalized_name.split())
    if len(name_words) > 0:
        for mapped_name, code in food_code_map.items():
            mapped_words = set(normalize_food_name(mapped_name).split())
            # If most words match, it's likely a match
            common_words = name_words & mapped_words
            if len(common_words) > 0 and len(common_words) / len(name_words) > 0.6:
                return code
    
    return None

# Step 2: Process each sheet
sheets_to_process = {
    'Food Items': 1,  # Column with food name
    'Nutritional Status': 1,  # Column with food name (Foods)
    'Carbon Footprint (SuEatable)': 2,  # Column with food name (Food commodity ITEM)
    'Water Footprint (SuEatable)': 2,  # Column with food name (Food commodity ITEM)
}

for sheet_name, food_col in sheets_to_process.items():
    if sheet_name not in wb.sheetnames:
        continue
    
    sheet = wb[sheet_name]
    print(f"Processing {sheet_name}...")
    
    # Check if food_code column already exists
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    
    if 'food_code' in headers:
        print(f"  - food_code column already exists, skipping\n")
        continue
    
    # Add food_code header
    new_col = sheet.max_column + 1
    sheet.cell(1, new_col).value = 'food_code'
    
    matched = 0
    unmatched = 0
    unmatched_samples = []
    
    # Match and add food codes
    for row in range(2, sheet.max_row + 1):
        food_name = sheet.cell(row, food_col).value
        code = quick_match(food_name, food_code_map)
        
        if code:
            sheet.cell(row, new_col).value = code
            matched += 1
        else:
            unmatched += 1
            if len(unmatched_samples) < 5:
                unmatched_samples.append(food_name)
    
    print(f"  ✓ Added food_code to {matched} rows ({unmatched} unmatched)")
    if unmatched_samples:
        print(f"  - Unmatched examples: {unmatched_samples}\n")
    else:
        print()

# Step 3: Save the updated workbook
output_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT)_CODED.xlsx'
wb.save(output_path)
print(f"✓ Saved updated file to: {output_path}")
