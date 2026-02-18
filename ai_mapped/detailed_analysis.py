import openpyxl

file_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT).xlsx'
wb = openpyxl.load_workbook(file_path)

print("=" * 80)
print("DETAILED DATA QUALITY ANALYSIS")
print("=" * 80)

# 1. Food Items - Typo and empty source column
print("\n1. FOOD ITEMS SHEET:")
sheet = wb['Food Items']
print(f"   Header Row: {[sheet.cell(1, c).value for c in range(1, 4)]}")
print(f"   - TYPO: 'Catergory' should be 'Category'")

source_filled = 0
for row in range(2, min(sheet.max_row + 1, 326)):
    if sheet.cell(row, 3).value:
        source_filled += 1
print(f"   - 'Source' column: Only {source_filled}/{sheet.max_row-1} rows have data (mostly empty)")

# 2. Nutritional Status - Check Amount column
print("\n2. NUTRITIONAL STATUS SHEET:")
sheet = wb['Nutritional Status']
headers = [sheet.cell(1, c).value for c in range(1, 10)]
print(f"   Headers: {headers}")
amount_filled = 0
for row in range(2, min(sheet.max_row + 1, 100)):
    if sheet.cell(row, 2).value:
        amount_filled += 1
print(f"   - 'Amount' column (col 2): Only {amount_filled}/98 sample rows have data")

# 3. Carbon Footprint - Check structure
print("\n3. CARBON FOOTPRINT (SuEatable) SHEET:")
sheet = wb['Carbon Footprint (SuEatable)']
headers = [sheet.cell(1, c).value for c in range(1, 10)]
print(f"   First 9 headers: {headers}")

# 4. DataField Carbon Footprint - Check structure
print("\n4. DATAFIELD CARBON FOOTPRINT SHEET:")
sheet = wb['DataField Carbon Footprint']
non_empty_headers = []
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(1, col).value
    if val:
        non_empty_headers.append((col, val))
print(f"   Note: Only {len(non_empty_headers)} out of {sheet.max_column} columns have headers")
print(f"   First few non-empty headers: {non_empty_headers[:5]}")
if non_empty_headers:
    print(f"   First non-empty header is in column {non_empty_headers[0][0]}: '{non_empty_headers[0][1]}'")

# 5. PIVOT table GHG - Check structure
print("\n5. PIVOT TABLE GHG DATAFIELD SHEET:")
sheet = wb['PIVOT table GHG DataField']
non_empty_headers = []
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(1, col).value
    if val:
        non_empty_headers.append((col, val))
print(f"   Note: Only {len(non_empty_headers)} out of {sheet.max_column} columns have headers")
print(f"   First few non-empty headers: {non_empty_headers[:5]}")

# 6. Science Data ALL - Check structure
print("\n6. SCIENCE DATA ALL SHEET:")
sheet = wb['Science Data ALL']
non_empty_headers = []
for col in range(1, sheet.max_column + 1):
    val = sheet.cell(1, col).value
    if val:
        non_empty_headers.append((col, val))
print(f"   Note: Only {len(non_empty_headers)} out of {sheet.max_column} columns have headers")
print(f"   Apparent structure issue: Headers don't exist or are irregularly placed")

# Check for data alignment issues
print("\n7. DATA ALIGNMENT ISSUES:")
print("   - Several sheets seem to have data but missing proper column headers")
print("   - This suggests the data might be raw/unprocessed or improperly formatted")

print("\n" + "=" * 80)
