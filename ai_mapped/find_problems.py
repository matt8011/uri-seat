import openpyxl
from collections import Counter

file_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT).xlsx'
wb = openpyxl.load_workbook(file_path)

print("=" * 80)
print("DATA QUALITY ISSUES FOUND")
print("=" * 80)

issues_found = False

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_issues = []
    
    # Check for empty cells in header row
    headers = []
    for col in range(1, sheet.max_column + 1):
        cell = sheet.cell(1, col)
        headers.append(cell.value)
        if cell.value is None:
            sheet_issues.append(f"  - Empty header in column {col}")
    
    # Check for duplicate headers
    header_counts = Counter(headers)
    for header, count in header_counts.items():
        if count > 1 and header is not None:
            sheet_issues.append(f"  - Duplicate header: '{header}' appears {count} times")
    
    # Check for typos in common words
    all_headers_str = ' '.join(str(h) for h in headers if h)
    if 'Catergory' in all_headers_str:
        sheet_issues.append(f"  - TYPO in header: 'Catergory' should be 'Category'")
    if 'UTILIZED IMPACT FACTORS' in all_headers_str and len(headers) > 1:
        # Check if this is a problematic header
        if headers[0] == 'UTILIZED IMPACT FACTORS':
            sheet_issues.append(f"  - Unusual header format: First cell is 'UTILIZED IMPACT FACTORS'")
    
    # Check for mostly empty columns
    for col in range(1, sheet.max_column + 1):
        non_empty = 0
        for row in range(2, min(sheet.max_row + 1, 100)):  # Sample first 100 rows
            if sheet.cell(row, col).value is not None:
                non_empty += 1
        if non_empty == 0:
            header = headers[col - 1]
            sheet_issues.append(f"  - Potentially empty column {col}: '{header}'")
    
    # Check for missing values in key columns
    if sheet_name == 'Food Items':
        for row in range(2, sheet.max_row + 1):
            name = sheet.cell(row, 1).value
            category = sheet.cell(row, 2).value
            if name is None or category is None:
                sheet_issues.append(f"  - Row {row}: Missing data (Name or Category)")
                break
    
    # Display issues for this sheet
    if sheet_issues:
        print(f"\n🔴 {sheet_name}:")
        for issue in sheet_issues[:5]:  # Show top 5 issues per sheet
            print(issue)
        if len(sheet_issues) > 5:
            print(f"  ... and {len(sheet_issues) - 5} more issues")
        issues_found = True

if not issues_found:
    print("\n✅ No obvious data quality issues found")

print("\n" + "=" * 80)
