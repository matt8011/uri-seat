import openpyxl
import re
from collections import defaultdict

file_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT).xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

# Step 1: Build efficient lookup from Affordability sheet
affordability_sheet = wb['Affordability']
food_code_map = {}  # Maps normalized name to code
word_to_codes = defaultdict(list)  # Maps words to possible codes

print("Building food code index...")
for row in range(2, min(affordability_sheet.max_row + 1, 5000)):  # Limit to first 5000 for efficiency
    code = affordability_sheet.cell(row, 1).value
    description = affordability_sheet.cell(row, 2).value
    if code and description:
        desc = str(description).lower().strip()
        food_code_map[desc] = code
        # Index by key words for faster matching
        words = re.findall(r'\w+', desc)
        for word in words:
            if len(word) > 3:  # Only index words longer than 3 chars
                word_to_codes[word].append(code)

print(f"Built index with {len(food_code_map)} foods and {len(word_to_codes)} keywords\n")

def smart_match(food_name, food_code_map, word_to_codes):
    """Fast matching using word-based indexing"""
    if not food_name:
        return None
    
    food_lower = str(food_name).lower().strip()
    
    # Exact match first (fast)
    if food_lower in food_code_map:
        return food_code_map[food_lower]
    
    # Extract key words
    words = [w for w in re.findall(r'\w+', food_lower) if len(w) > 3]
    if not words:
        return None
    
    # Use word index to find candidates
    candidates = defaultdict(int)
    for word in words:
        for code in word_to_codes.get(word, []):
            candidates[code] += 1
    
    # Return the code with most matching words
    if candidates:
        best_code = max(candidates, key=candidates.get)
        if candidates[best_code] >= len(words) * 0.5:  # At least 50% word match
            return best_code
    
    return None

# Step 2: Reload workbook without data_only for editing
wb = openpyxl.load_workbook(file_path)

sheets_to_process = {
    'Food Items': 1,
    'Nutritional Status': 1,
    'Carbon Footprint (SuEatable)': 2,
    'Water Footprint (SuEatable)': 2,
}

total_matched = 0
total_unmatched = 0

for sheet_name, food_col in sheets_to_process.items():
    if sheet_name not in wb.sheetnames:
        continue
    
    sheet = wb[sheet_name]
    print(f"Processing {sheet_name}...")
    
    # Check if food_code column already exists
    headers = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    if 'food_code' in headers:
        print(f"  - Column already exists\n")
        continue
    
    # Add food_code header
    new_col = sheet.max_column + 1
    sheet.cell(1, new_col).value = 'food_code'
    
    matched = 0
    unmatched = 0
    samples = []
    
    # Process all rows
    max_rows = sheet.max_row
    for row in range(2, max_rows + 1):
        food_name = sheet.cell(row, food_col).value
        code = smart_match(food_name, food_code_map, word_to_codes)
        
        if code:
            sheet.cell(row, new_col).value = code
            matched += 1
        else:
            unmatched += 1
            if len(samples) < 3:
                samples.append(food_name)
        
        # Progress indicator for large sheets
        if row % 1000 == 0:
            print(f"  Processing row {row}/{max_rows}...")
    
    print(f"  ✓ Matched: {matched}, Unmatched: {unmatched}")
    if samples:
        print(f"  - Samples: {samples}\n")
    else:
        print()
    
    total_matched += matched
    total_unmatched += unmatched

# Step 3: Save
output_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT)_CODED.xlsx'
wb.save(output_path)
print(f"\n✓ COMPLETE! Saved: {output_path}")
print(f"✓ Total matched: {total_matched}, unmatched: {total_unmatched}")
