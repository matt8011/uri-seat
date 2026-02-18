import openpyxl
from pathlib import Path

# Load the workbook
file_path = r'c:\Users\gyank\Desktop\Research\SEAT\Sustainable Eating Assessment Tool (SEAT).xlsx'
wb = openpyxl.load_workbook(file_path)

print("=" * 80)
print("EXCEL FILE ANALYSIS: Sustainable Eating Assessment Tool (SEAT)")
print("=" * 80)
print(f"\nTotal Sheets: {len(wb.sheetnames)}\n")

for i, sheet_name in enumerate(wb.sheetnames, 1):
    sheet = wb[sheet_name]
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    print(f"\n{i}. Sheet: '{sheet_name}'")
    print(f"   - Dimensions: {max_row} rows × {max_col} columns")
    
    # Get header row if present
    if max_row > 0:
        headers = []
        for col in range(1, min(max_col + 1, 10)):  # Show first 10 columns
            cell = sheet.cell(1, col)
            if cell.value:
                headers.append(str(cell.value))
        if headers:
            print(f"   - First few headers: {', '.join(headers)}")

print("\n" + "=" * 80)
